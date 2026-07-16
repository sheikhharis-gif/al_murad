import datetime
import math
from pathlib import Path
from django.core.management.base import BaseCommand
from openpyxl import load_workbook
from django.utils import timezone
from decimal import Decimal

from masters.models import City, Route, Client, Vendor, Vehicle, Expense
from operations.models import Job, Trip

class Command(BaseCommand):
    help = "Import data from AMG.xlsx into database"

    def handle(self, *args, **options):
        self.stdout.write("Starting import from AMG.xlsx...")
        
        file_path = Path("AMG.xlsx")
        if not file_path.exists():
            self.stdout.write(self.style.ERROR("AMG.xlsx not found in project root."))
            return

        # Load workbook
        wb = load_workbook(file_path, data_only=True)
        if "Data Sheet" not in wb.sheetnames:
            self.stdout.write(self.style.ERROR("Sheet 'Data Sheet' not found in AMG.xlsx"))
            return

        sheet = wb["Data Sheet"]
        
        # We need a default vendor for Vehicles
        default_vendor, _ = Vendor.objects.get_or_create(
            name="Default Vendor",
            defaults={
                "phone": "03001234567",
                "poc": "Default POC",
                "ntn": "VD-99999",
                "address": "Karachi Head Office"
            }
        )

        row_count = 0
        success_count = 0
        skipped_count = 0

        # Read starting from row 2
        for idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            # Check if row is empty
            if not any(cell is not None for cell in row):
                continue
                
            row_count += 1
            
            # Map columns by indexing (0-based)
            # From amg_summary:
            # 0: Profit
            # 1: P&L %
            # 2: Customer (code/short name)
            # 3: Customer Name
            # 4: Order Reference
            # 5: Origin
            # 6: Destination
            # 7: Weight (Tons)
            # 8: Departure
            # 9: Arrival
            # 10: Delivery
            # 11: Vehicle #
            # 12: Vehicle Type
            # 13: Route
            # 14: Bilty #
            # 15: Status
            # 16: KMs
            # 17: Fuel Average
            # 18: Fuel in Liters
            # 19: Fuel Price
            # 20: Fuel Amount
            # 21: Toll Tax
            # 22: Incentive
            # 23: Food
            # 24: Police
            # 25: Card
            # 26: Maintenance
            # 27: Other
            # 28: Total Expense
            # 29: Trip Charges
            # 30: Additional Stop
            # 31: Detention
            # 32: Labor Charges
            # 33: Total Charges

            # Stop if customer name is empty or error cells
            cust_name = row[3] or row[2]
            if not cust_name or str(cust_name).startswith("#") or cust_name == "-":
                skipped_count += 1
                continue

            # Parse dates
            def parse_date(val):
                if isinstance(val, (datetime.datetime, datetime.date)):
                    return val
                if isinstance(val, str):
                    for fmt in ("%Y-%m-%dT%H:%M:%S", "%Y-%m-%d %H:%M:%S", "%Y-%m-%d", "%d-%m-%Y"):
                        try:
                            return datetime.datetime.strptime(val, fmt).date()
                        except ValueError:
                            pass
                return datetime.date.today()

            trip_date = parse_date(row[8] or row[9] or row[10])

            # 1. Client
            client_name = str(cust_name).strip()
            client_code = str(row[2]).strip() if row[2] else client_name[:10]
            # Create a clean unique NTN
            ntn_val = f"CL-{hash(client_name) % 100000000:08d}"
            
            client, _ = Client.objects.get_or_create(
                name=client_name,
                defaults={
                    "poc": "Imported POC",
                    "ntn": ntn_val,
                    "address": "Imported Address",
                    "is_active": True
                }
            )

            # 2. Cities & Route
            origin_name = str(row[5]).strip() if row[5] else "Karachi"
            dest_name = str(row[6]).strip() if row[6] else "Lahore"
            
            def get_or_create_city(name):
                code = "".join([c for c in name if c.isalnum()]).upper()[:10]
                if not code:
                    code = "CITY"
                city, _ = City.objects.get_or_create(
                    code=code,
                    defaults={
                        "name": name,
                        "latitude": Decimal("24.8607"),
                        "longitude": Decimal("67.0011")
                    }
                )
                return city

            origin_city = get_or_create_city(origin_name)
            dest_city = get_or_create_city(dest_name)

            try:
                distance = int(row[16]) if row[16] and not str(row[16]).startswith("#") else 1200
            except:
                distance = 1200

            route, _ = Route.objects.get_or_create(
                origin=origin_city,
                destination=dest_city,
                defaults={"distance_km": distance}
            )

            # 3. Vehicle
            veh_num = str(row[11]).strip() if row[11] else "JV-TEMP"
            if not veh_num or veh_num == "-" or str(veh_num).startswith("#"):
                veh_num = "JV-TEMP"
                
            vehicle_type_raw = str(row[12]).strip() if row[12] else "40ft Dry"
            # Map type to choice if possible
            vehicle_type = "40ft Dry"
            for choice in [c[0] for c in Vehicle.VEHICLE_TYPES]:
                if vehicle_type_raw.lower() in choice.lower():
                    vehicle_type = choice
                    break

            vehicle, _ = Vehicle.objects.get_or_create(
                vehicle_number=veh_num,
                defaults={
                    "vendor": default_vendor,
                    "vehicle_mode": "OWN",
                    "vehicle_type": vehicle_type,
                    "wheeler": "2x4",
                    "current_location": origin_name,
                    "is_active": True,
                    "current_km": distance
                }
            )

            # 4. Job
            # Status mapping
            status_raw = str(row[15]).strip().lower() if row[15] else "completed"
            status = "completed"
            if "progress" in status_raw:
                status = "in_progress"
            elif "return" in status_raw:
                status = "returning"
            elif "cancel" in status_raw:
                status = "cancelled"

            # Check if there is a job already on this day or create a new job per row
            job = Job.objects.create(
                vehicle=vehicle,
                status=status,
                job_date=trip_date,
                remarks=f"Imported Job from sheet row {idx}"
            )

            # 5. Trip
            bilty = str(row[14]).strip() if row[14] else f"B-{job.job_number}"
            
            def clean_float(val, default=0.0):
                if val is None or str(val).startswith("#") or val == "-":
                    return default
                try:
                    return float(val)
                except:
                    return default

            weight = clean_float(row[7], 1.0)
            trip_charges = clean_float(row[29], 0.0)
            detention = clean_float(row[31], 0.0)
            
            # calculate rate so weight * rate = trip_charges
            rate = trip_charges / weight if weight > 0 else trip_charges

            trip = Trip.objects.create(
                job=job,
                client=client,
                vehicle=vehicle,
                trip_date=trip_date,
                route=route,
                bilty_number=bilty,
                weight=Decimal(str(weight)),
                rate=Decimal(str(rate)),
                detention=Decimal(str(detention))
            )

            # 6. Expense
            fuel_liters = clean_float(row[18], 0.0)
            fuel_price = clean_float(row[19], 0.0)
            fuel_amount = clean_float(row[20], 0.0)
            toll_tax = clean_float(row[21], 0.0)
            inam = clean_float(row[22], 0.0)
            police = clean_float(row[24], 0.0)
            food = clean_float(row[23], 0.0)
            card = clean_float(row[25], 0.0)
            maint = clean_float(row[26], 0.0)
            other = clean_float(row[27], 0.0)

            Expense.objects.create(
                trip=trip,
                date=trip_date,
                fuel_liter=fuel_liters,
                fuel_rate=fuel_price,
                fuel_amount=fuel_amount,
                toll_tax=toll_tax,
                inam=inam,
                police=police,
                food=food,
                card=card,
                maintenance=maint,
                other=other,
                slip_no=f"SLIP-{idx}"
            )

            success_count += 1

        self.stdout.write(self.style.SUCCESS(
            f"Successfully processed {row_count} rows. "
            f"Created {success_count} jobs/trips. Skipped {skipped_count} rows."
        ))
