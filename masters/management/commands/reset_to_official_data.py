from pathlib import Path

from django.conf import settings
from django.core.management import call_command
from django.core.management.base import BaseCommand
from openpyxl import load_workbook

from masters.models import Client, Driver, Vehicle, Vendor
from operations.models import Invoice, Job, Trip


def clean(val):
    if val is None:
        return ""
    return str(val).strip()


class Command(BaseCommand):
    help = (
        "Wipes all Vehicle/Driver/Client records not present in AMG Details.xlsx "
        "(Vehicle Data / Staff & Drivers / Customer sheets) along with every Job/Trip/"
        "Invoice, then re-runs import_amg_details so the database holds exactly the "
        "official dataset. Safe to re-run."
    )

    def add_arguments(self, parser):
        parser.add_argument(
            "path",
            nargs="?",
            default="AMG Details.xlsx",
            help="Path to the details Excel file (relative to the project root)",
        )

    def handle(self, *args, **options):
        path = Path(settings.BASE_DIR) / options["path"]
        if not path.exists():
            self.stderr.write(self.style.ERROR(f"File not found: {path}"))
            return

        wb = load_workbook(path, data_only=True)

        official_plates = set()
        for row in wb["Vehicle Data"].iter_rows(min_row=4, values_only=True):
            plate = clean(row[2])
            if plate:
                official_plates.add(plate)

        official_cnics = set()
        for row in wb["Staff & Drivers"].iter_rows(min_row=4, values_only=True):
            if clean(row[3]) != "Driver":
                continue
            cnic = clean(row[9])
            if cnic:
                official_cnics.add(cnic)

        official_clients = set()
        for row in wb["Customer"].iter_rows(min_row=4, values_only=True):
            name = clean(row[1])
            if name:
                official_clients.add(name)

        trips_deleted = Trip.objects.all().count()
        Trip.objects.all().delete()
        invoices_deleted = Invoice.objects.all().count()
        Invoice.objects.all().delete()
        jobs_deleted = Job.objects.all().count()
        Job.objects.all().delete()

        vehicles_qs = Vehicle.objects.exclude(vehicle_number__in=official_plates)
        vehicles_deleted = vehicles_qs.count()
        vehicles_qs.delete()

        drivers_qs = Driver.objects.exclude(cnic__in=official_cnics)
        drivers_deleted = drivers_qs.count()
        drivers_qs.delete()

        clients_qs = Client.objects.exclude(name__in=official_clients)
        clients_deleted = clients_qs.count()
        clients_qs.delete()

        vendors_deleted = 0
        for vendor in Vendor.objects.exclude(name="Not Specified"):
            if vendor.vehicles.count() == 0:
                vendor.delete()
                vendors_deleted += 1

        self.stdout.write(self.style.SUCCESS(
            f"Removed: {trips_deleted} trips, {invoices_deleted} invoices, {jobs_deleted} jobs, "
            f"{vehicles_deleted} non-official vehicles, {drivers_deleted} non-official drivers, "
            f"{clients_deleted} non-official clients, {vendors_deleted} orphaned vendors."
        ))

        call_command("import_amg_details", options["path"])
