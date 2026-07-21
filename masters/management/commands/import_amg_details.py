import datetime
import re
from pathlib import Path

from django.conf import settings
from django.core.management.base import BaseCommand
from openpyxl import load_workbook

from masters.models import Client, Driver, Vehicle, Vendor

PLACEHOLDER_EXPIRY = datetime.date(2030, 12, 31)
PLACEHOLDER_JOINING = datetime.date(2020, 1, 1)


def clean(val):
    if val is None:
        return ""
    s = str(val).strip()
    if s in ("-", "--", "---", "NOT DEFINE", "None"):
        return ""
    return s


def parse_date(val, default=None):
    if isinstance(val, (datetime.datetime, datetime.date)):
        return val if isinstance(val, datetime.date) and not isinstance(val, datetime.datetime) else val.date()
    return default


def normalize_vehicle_type(raw):
    s = clean(raw).upper()
    if not s:
        return "Mini Truck"
    if "MINI" in s:
        return "Mini Truck"
    m = re.match(r"(\d+)\s*FT\s*(DRY|REFFER|REEFER)", s)
    if m:
        num, body = m.group(1), m.group(2)
        body = "Reefer" if "REF" in body else "Dry"
        return f"{num}ft {body}"
    return "Mini Truck"


def normalize_wheeler(raw):
    s = clean(raw)
    if s in ("6", "14"):
        return s
    return "6"


class Command(BaseCommand):
    help = "Import vehicles, drivers and customers from 'AMG Details.xlsx' (Vehicle Data / Staff & Drivers / Customer sheets)"

    def add_arguments(self, parser):
        parser.add_argument(
            "path",
            nargs="?",
            default="AMG Details.xlsx",
            help="Path to the details Excel file (relative to the project root)",
        )

    def handle(self, *args, **options):
        path = Path(settings.BASE_DIR) / options["path"]
        if not path.exists():
            self.stderr.write(self.style.ERROR(f"File not found: {path}"))
            return

        wb = load_workbook(path, data_only=True)

        self.import_customers(wb)
        self.import_drivers(wb)
        self.import_vehicles(wb)

    def import_customers(self, wb):
        if "Customer" not in wb.sheetnames:
            self.stderr.write(self.style.WARNING("No 'Customer' sheet found, skipping."))
            return
        ws = wb["Customer"]

        created, existing, used_ntns = 0, 0, set()
        for row in ws.iter_rows(min_row=4, values_only=True):
            name = clean(row[1])
            if not name:
                continue

            ntn = clean(row[2])
            code = clean(row[3])
            client_type = clean(row[4])
            address = clean(row[5])
            poc = clean(row[6])
            billing_company = clean(row[7])

            if not ntn:
                ntn = f"NTN-{code or name[:10]}"
            if ntn in used_ntns:
                ntn = f"{ntn}-{code or name[:6]}"
            used_ntns.add(ntn)

            client, was_created = Client.objects.get_or_create(
                name=name,
                defaults={
                    "poc": poc,
                    "ntn": ntn,
                    "address": address or "Not Provided",
                    "billing_company": billing_company,
                    "is_active": True,
                },
            )
            if was_created:
                created += 1
            else:
                existing += 1

        self.stdout.write(self.style.SUCCESS(
            f"Customers -> created: {created}, already existed: {existing}"
        ))

    def import_drivers(self, wb):
        if "Staff & Drivers" not in wb.sheetnames:
            self.stderr.write(self.style.WARNING("No 'Staff & Drivers' sheet found, skipping."))
            return
        ws = wb["Staff & Drivers"]

        created, existing, skipped_non_driver = 0, 0, 0
        for row in ws.iter_rows(min_row=4, values_only=True):
            name = clean(row[4])
            designation = clean(row[3])
            if not name:
                continue
            if designation != "Driver":
                skipped_non_driver += 1
                continue

            father_name = clean(row[5]) or "Not Provided"
            mobile = clean(row[6]) or clean(row[7]) or "N/A"
            cnic = clean(row[9])
            if not cnic:
                continue

            license_number = clean(row[10]) or "N/A"
            address = clean(row[11]) or "Not Provided"
            cnic_expiry = parse_date(row[13], PLACEHOLDER_EXPIRY)
            license_expiry = parse_date(row[16], PLACEHOLDER_EXPIRY)
            ref1_name = clean(row[17])
            ref1_mobile = clean(row[18])
            ref2_name = clean(row[19])
            ref2_mobile = clean(row[20]) if len(row) > 20 else ""

            driver, was_created = Driver.objects.get_or_create(
                cnic=cnic,
                defaults={
                    "name": name,
                    "father_name": father_name,
                    "address": address,
                    "mobile": mobile,
                    "cnic_expiry": cnic_expiry,
                    "license_number": license_number,
                    "license_expiry": license_expiry,
                    "reference1_name": ref1_name,
                    "reference1_mobile": ref1_mobile,
                    "reference2_name": ref2_name,
                    "reference2_mobile": ref2_mobile,
                    "joining_date": PLACEHOLDER_JOINING,
                    "is_active": True,
                },
            )
            if was_created:
                created += 1
            else:
                existing += 1

        self.stdout.write(self.style.SUCCESS(
            f"Drivers -> created: {created}, already existed: {existing}, "
            f"non-driver staff skipped: {skipped_non_driver}"
        ))

    def import_vehicles(self, wb):
        if "Vehicle Data" not in wb.sheetnames:
            self.stderr.write(self.style.WARNING("No 'Vehicle Data' sheet found, skipping."))
            return
        ws = wb["Vehicle Data"]

        default_vendor, _ = Vendor.objects.get_or_create(
            name="Not Specified",
            defaults={
                "phone": "N/A",
                "poc": "N/A",
                "ntn": "",
                "address": "Not Specified",
            },
        )

        created, existing, used_plates = 0, 0, set()
        for row in ws.iter_rows(min_row=4, values_only=True):
            sr_no = row[1]
            if not sr_no:
                continue

            plate = clean(row[2])
            if not plate:
                continue
            if plate in used_plates:
                continue
            used_plates.add(plate)

            engine_no = clean(row[3])
            chassis_no = clean(row[4])
            color = clean(row[5])
            vehicle_type = normalize_vehicle_type(row[7])
            mode_raw = clean(row[8]).upper()
            vehicle_mode = "OWN" if mode_raw == "OWN" else "RENTAL" if mode_raw else "OWN"
            wheeler = normalize_wheeler(row[9])

            vehicle, was_created = Vehicle.objects.get_or_create(
                vehicle_number=plate,
                defaults={
                    "vendor": default_vendor,
                    "vehicle_mode": vehicle_mode,
                    "vehicle_type": vehicle_type,
                    "engine_no": engine_no,
                    "chassis_no": chassis_no,
                    "color": color,
                    "wheeler": wheeler,
                    "is_active": True,
                    "current_km": 0,
                },
            )
            if was_created:
                created += 1
            else:
                existing += 1

        self.stdout.write(self.style.SUCCESS(
            f"Vehicles -> created: {created}, already existed: {existing}"
        ))
