import datetime
from pathlib import Path

from django.conf import settings
from django.core.management.base import BaseCommand
from openpyxl import load_workbook

from masters.models import Driver, DriverSalary

PLACEHOLDER_EXPIRY = datetime.date(2030, 12, 31)
PLACEHOLDER_JOINING = datetime.date(2020, 1, 1)


class Command(BaseCommand):
    help = "Import driver salary records from the 'Salary Engine' sheet of a salary workbook (e.g. 'JUN-26 Salary Sheet.xlsx')"

    def add_arguments(self, parser):
        parser.add_argument(
            "path",
            nargs="?",
            default="JUN-26 Salary Sheet.xlsx",
            help="Path to the salary Excel file (relative to the project root)",
        )
        parser.add_argument(
            "--month",
            default="2026-06-01",
            help="Salary month as YYYY-MM-DD (defaults to 2026-06-01, matching the JUN-26 sheet)",
        )

    def handle(self, *args, **options):
        path = Path(settings.BASE_DIR) / options["path"]
        if not path.exists():
            self.stderr.write(self.style.ERROR(f"File not found: {path}"))
            return

        salary_month = datetime.date.fromisoformat(options["month"])

        wb = load_workbook(path, data_only=True)
        if "Salary Engine" not in wb.sheetnames:
            self.stderr.write(self.style.ERROR("This workbook has no 'Salary Engine' sheet"))
            return
        ws = wb["Salary Engine"]

        created_drivers = 0
        existing_drivers = 0
        created_salaries = 0
        updated_salaries = 0
        skipped_blank = 0

        def num(val):
            if val is None or str(val).startswith("#"):
                return 0
            try:
                return float(val)
            except (TypeError, ValueError):
                return 0

        for row in ws.iter_rows(min_row=3, values_only=True):
            emp_id, name, designation = row[0], row[1], row[2]
            if not emp_id:
                continue
            if not name or designation != "Driver":
                skipped_blank += 1
                continue

            (
                present_days, absent_days, sundays,
                base_salary, per_day_rate, earned_base_salary,
                attendance_allowance, total_gross_salary,
                previous_advance, new_advance_taken, advance_deduction,
                net_payable_salary, status,
            ) = row[3], row[4], row[5], row[6], row[7], row[8], row[9], row[10], row[11], row[12], row[13], row[14], row[15]

            driver, created = Driver.objects.get_or_create(
                cnic=f"TEST-{emp_id}",
                defaults={
                    "name": name,
                    "father_name": "Not Provided",
                    "address": "Not Provided (imported from salary sheet)",
                    "mobile": "0300-0000000",
                    "cnic_expiry": PLACEHOLDER_EXPIRY,
                    "license_number": f"TEST-LIC-{emp_id}",
                    "license_expiry": PLACEHOLDER_EXPIRY,
                    "joining_date": PLACEHOLDER_JOINING,
                    "is_active": True,
                },
            )
            if created:
                created_drivers += 1
            else:
                existing_drivers += 1
                if not driver.employee_id:
                    driver.save()  # backfill Employee ID for drivers created before that field existed

            salary_fields = {
                "emp_id": emp_id,
                "designation": designation,
                "present_days": int(num(present_days)),
                "absent_days": int(num(absent_days)),
                "sundays": int(num(sundays)),
                "base_salary": num(base_salary),
                "per_day_rate": num(per_day_rate),
                "earned_base_salary": num(earned_base_salary),
                "attendance_allowance": num(attendance_allowance),
                "total_gross_salary": num(total_gross_salary),
                "previous_advance": num(previous_advance),
                "new_advance_taken": num(new_advance_taken),
                "advance_deduction": num(advance_deduction),
                "net_payable_salary": num(net_payable_salary),
                "status": status or "Active",
            }

            salary, s_created = DriverSalary.objects.get_or_create(
                driver=driver,
                month=salary_month,
                defaults={**salary_fields, "paid": False},
            )
            if s_created:
                created_salaries += 1
            else:
                # Refresh sheet-derived fields on records created before this data was
                # available (e.g. by an older version of this command), without
                # touching "paid" since that's a manual toggle the user may have set.
                for field, value in salary_fields.items():
                    setattr(salary, field, value)
                salary.save()
                updated_salaries += 1

        self.stdout.write(self.style.SUCCESS(
            f"Drivers created: {created_drivers}, already existed: {existing_drivers}, "
            f"salary records created: {created_salaries}, updated: {updated_salaries}, "
            f"blank/non-driver rows skipped: {skipped_blank}"
        ))
