import datetime as dt
import io
import json
from datetime import date
from decimal import Decimal
from pathlib import Path
from django.contrib.auth.decorators import login_required
from django.db.models import Count, Q, Sum
from django.db.models.functions import TruncMonth
from django.http import HttpResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.template.loader import get_template
from xhtml2pdf import pisa

from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.platypus import Image as RLImage, Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

from .forms import JobForm, TripForm
from .models import Job, Trip

from masters.models import (
    City,
    Client,
    Driver,
    Expense,
    Route,
    Vendor,
    Vehicle,
    VehicleMaintenance,
)

# ================= DASHBOARD =================
@login_required
def dashboard(request):
    # --- LOGIC FOR EXPIRY ALERTS ---
    today = dt.date.today()
    warning_limit = today + dt.timedelta(days=15)

    # Woh gaariyan nikaalo jinka koi bhi permit 15 din mein expire ho raha ho
    critical_vehicles = Vehicle.objects.filter(
        Q(sindh_permit_expiry__lte=warning_limit) |
        Q(punjab_permit_expiry__lte=warning_limit) |
        Q(kpk_permit_expiry__lte=warning_limit) |
        Q(balochistan_permit_expiry__lte=warning_limit) |
        Q(fitness_expiry_sindh__lte=warning_limit) |
        Q(fitness_expiry_punjab__lte=warning_limit) |
        Q(fitness_expiry_kpk__lte=warning_limit) |
        Q(fitness_expiry_balochistan__lte=warning_limit)
    ).distinct()

    # Baqi calculations
    total_trips = Trip.objects.count()
    total_vehicles = Vehicle.objects.count()
    total_drivers = Driver.objects.count()
    total_freight = Trip.objects.aggregate(total=Sum("freight"))["total"] or 0
    total_expense = Expense.objects.aggregate(total=Sum("total_expense"))["total"] or 0
    profit = total_freight - Decimal(total_expense)

    upcoming_maintenance = VehicleMaintenance.objects.filter(
        next_due_date__isnull=False
    ).order_by("next_due_date")[:8]

    six_months_ago = today - dt.timedelta(days=180)
    monthly_trips = (
        Trip.objects.filter(trip_date__gte=six_months_ago)
        .annotate(month=TruncMonth("trip_date"))
        .values("month")
        .annotate(count=Count("id"))
        .order_by("month")
    )
    chart_labels = [entry["month"].strftime("%b %Y") for entry in monthly_trips]
    chart_trips = [entry["count"] for entry in monthly_trips]

    amg_sheets = []
    selected_sheet = request.GET.get("sheet", "")
    amg_headers = []
    amg_rows = []
    amg_row_count = 0
    amg_column_count = 0
    amg_error = None

    workbook_path = Path(__file__).resolve().parent.parent / "AMG.xlsx"
    if workbook_path.exists():
        try:
            workbook = load_workbook(workbook_path, data_only=True)
            amg_sheets = workbook.sheetnames
            if selected_sheet not in amg_sheets:
                selected_sheet = amg_sheets[0] if amg_sheets else ""
            if selected_sheet:
                worksheet = workbook[selected_sheet]
                raw_rows = list(worksheet.iter_rows(values_only=True))
                if raw_rows:
                    amg_headers = [
                        str(cell).replace("\n", " ").strip() if cell is not None else ""
                        for cell in raw_rows[0]
                    ]
                    for row in raw_rows[1:]:
                        if not any(cell is not None and str(cell).strip() for cell in row):
                            continue
                        amg_rows.append([_normalize_amg_cell(cell) for cell in row])
                    amg_row_count = len(amg_rows)
                    amg_column_count = len(amg_headers)
        except Exception as exc:
            amg_error = str(exc)
    else:
        amg_error = "AMG.xlsx file not found in project root."

    context = {
        "critical_vehicles": critical_vehicles,
        "today": today,
        "warning_date": warning_limit,
        "total_trips": total_trips,
        "total_vehicles": total_vehicles,
        "total_drivers": total_drivers,
        "total_freight": total_freight,
        "total_expense": total_expense,
        "profit": profit,
        "upcoming_maintenance": upcoming_maintenance,
        "chart_labels": json.dumps(chart_labels),
        "chart_trips": json.dumps(chart_trips),
        "amg_sheets": amg_sheets,
        "selected_sheet": selected_sheet,
        "amg_headers": amg_headers,
        "amg_rows": amg_rows,
        "amg_row_count": amg_row_count,
        "amg_column_count": amg_column_count,
        "amg_error": amg_error,
    }
    return render(request, "dashboard.html", context)


def _normalize_amg_cell(value):
    if value is None:
        return ""
    if isinstance(value, dt.datetime):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, dt.date):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, dt.time):
        return value.strftime("%H:%M:%S")
    if isinstance(value, float) and value.is_integer():
        return int(value)
    return value

# ================= JOBS =================
def job_list(request):
    jobs = Job.objects.all().order_by("-job_date")
    return render(request, "operations/job_list.html", {"jobs": jobs})

def job_add(request):
    form = JobForm(request.POST or None)
    if request.method == "POST" and form.is_valid():
        form.save()
        return redirect("job_list")
    return render(request, "operations/job_form.html", {"form": form})

def job_edit(request, job_id):
    # ERROR FIX: id=job_id ko job_number=job_id kar diya
    job = get_object_or_404(Job, job_number=job_id)
    form = JobForm(request.POST or None, instance=job)
    if request.method == "POST" and form.is_valid():
        form.save()
        return redirect("job_list")
    return render(request, "operations/job_form.html", {"form": form, "job": job})

def job_delete(request, job_id):
    # ERROR FIX: id=job_id ko job_number=job_id kar diya
    job = get_object_or_404(Job, job_number=job_id)
    job.delete()
    return redirect("job_list")

# ================= TRIPS =================

def _get_report_context(request):
    report = request.GET.get("report", "own")
    selected_client = request.GET.get("client")
    selected_job = request.GET.get("job")
    selected_vehicle = request.GET.get("vehicle")
    selected_route = request.GET.get("route")
    start_date = request.GET.get("start_date")
    end_date = request.GET.get("end_date")

    trips = Trip.objects.all().select_related("job", "client", "vehicle__driver", "route").order_by("-trip_date")

    # Primary Mode filter
    if report == "own":
        trips = trips.filter(vehicle__vehicle_mode="OWN")
        report_label = "OWN Trips"
    else:
        report_label = "All Trips"

    # Multi-criteria refinement
    if selected_client:
        trips = trips.filter(client_id=selected_client)
        report_label = f"Filtered Trips - Client"
    if selected_job:
        trips = trips.filter(job_id=selected_job)
        report_label = f"Filtered Trips - Job #{selected_job}"
    if selected_vehicle:
        trips = trips.filter(vehicle_id=selected_vehicle)
        report_label = f"Filtered Trips - Vehicle"
    if selected_route:
        trips = trips.filter(route_id=selected_route)
        report_label = f"Filtered Trips - Route"
    if start_date:
        trips = trips.filter(trip_date__gte=start_date)
    if end_date:
        trips = trips.filter(trip_date__lte=end_date)

    return {
        "trips": trips,
        "report": report,
        "report_label": report_label,
        "selected_client": selected_client,
        "selected_job": selected_job,
        "selected_vehicle": selected_vehicle,
        "selected_route": selected_route,
        "start_date": start_date,
        "end_date": end_date,
    }


def trip_list(request):
    report_context = _get_report_context(request)
    trips = report_context["trips"]
    
    # Financial & Operational Metrics
    total_freight = trips.aggregate(total=Sum("freight"))["total"] or 0
    
    trip_ids = trips.values_list('id', flat=True)
    expenses_aggregates = Expense.objects.filter(trip_id__in=trip_ids).aggregate(
        total_exp=Sum('total_expense'),
        total_liters=Sum('fuel_liter')
    )
    total_expense = expenses_aggregates['total_exp'] or 0
    total_liters = expenses_aggregates['total_liters'] or 0
    net_profit = Decimal(str(total_freight)) - Decimal(str(total_expense))
    pl_percentage = (net_profit / Decimal(str(total_freight)) * 100) if total_freight > 0 else 0

    # Per-trip Total = Freight - that trip's own expenses
    expense_by_trip = {
        row["trip_id"]: row["total"]
        for row in Expense.objects.filter(trip_id__in=trip_ids)
        .values("trip_id")
        .annotate(total=Sum("total_expense"))
    }
    for t in trips:
        t.net_total = t.freight - Decimal(str(expense_by_trip.get(t.id, 0)))

    clients = Client.objects.filter(is_active=True).order_by("name")
    jobs = Job.objects.all().order_by("-job_date")
    vehicles = Vehicle.objects.all().order_by("vehicle_number")
    routes = Route.objects.all().order_by("route_code")

    context = {
        "trips": trips,
        "total_freight": total_freight,
        "total_expense": total_expense,
        "total_liters": total_liters,
        "net_profit": net_profit,
        "pl_percentage": pl_percentage,
        "clients": clients,
        "jobs": jobs,
        "vehicles": vehicles,
        "routes": routes,
        **report_context,
    }
    return render(request, "operations/trip_list.html", context)


@login_required
def trips_excel(request):
    report_context = _get_report_context(request)
    trips = list(report_context["trips"].order_by("trip_date"))

    wb = Workbook()
    ws = wb.active
    ws.title = "Trip Sheet Own"

    def set_cell(row, col, value, bold=False, align='left'):
        cell = ws.cell(row=row, column=col, value=value)
        if bold:
            cell.font = Font(bold=True)
        if align:
            cell.alignment = Alignment(horizontal=align)
        return cell

    ws.merge_cells('C2:H2')
    set_cell(2, 3, 'Trip Sheet Own', bold=True, align='center')

    set_cell(6, 3, 'Job #', bold=True)
    set_cell(6, 5, 'Vehicle Number', bold=True)
    set_cell(6, 7, 'Driver 1 Name', bold=True)
    set_cell(7, 7, 'Driver 2 Name', bold=True)
    set_cell(9, 3, 'Meter Out', bold=True)
    set_cell(9, 5, 'Meter In', bold=True)
    set_cell(9, 7, 'KMs', bold=True)
    set_cell(10, 7, 'Fuel in Liters', bold=True)
    set_cell(11, 7, 'Fuel Average', bold=True)
    set_cell(27, 3, 'Expense', bold=True)
    set_cell(28, 3, 'Toll Tax', bold=True)
    set_cell(29, 3, 'Incentive', bold=True)
    set_cell(30, 3, 'Food', bold=True)
    set_cell(31, 3, 'Police', bold=True)
    set_cell(32, 3, 'Card', bold=True)
    set_cell(33, 3, 'Maintenance', bold=True)
    set_cell(34, 3, 'Other', bold=True)
    set_cell(35, 3, 'Rental', bold=True)
    set_cell(36, 3, 'Detention', bold=True)
    set_cell(37, 3, 'Total', bold=True)
    set_cell(39, 3, 'Fuel', bold=True)
    set_cell(40, 3, 'Date', bold=True)
    set_cell(40, 4, 'Pump Name', bold=True)
    set_cell(40, 5, 'Slip #', bold=True)
    set_cell(40, 6, 'Fuel In Liters', bold=True)
    set_cell(40, 7, 'Fuel Rate', bold=True)
    set_cell(40, 8, 'Amount', bold=True)

    if trips:
        first_trip = trips[0]
        set_cell(6, 4, first_trip.job.job_number)
        set_cell(6, 6, first_trip.vehicle.vehicle_number)
        set_cell(6, 8, first_trip.vehicle.driver.name if first_trip.vehicle.driver else '')
        set_cell(7, 8, '')

    for idx in range(3):
        label_col = 3 + idx * 2
        value_col = label_col + 1
        if idx < len(trips):
            trip = trips[idx]
            set_cell(13, label_col, 'Trip #', bold=True)
            set_cell(13, value_col, trip.trip_no)
            set_cell(14, label_col, 'Bilty #', bold=True)
            set_cell(14, value_col, trip.bilty_number)
            set_cell(15, label_col, 'Client', bold=True)
            set_cell(15, value_col, trip.client.name)
            set_cell(16, label_col, 'Route', bold=True)
            set_cell(16, value_col, str(trip.route))
            set_cell(18, label_col, 'Freight', bold=True)
            set_cell(18, value_col, float(trip.freight))
            set_cell(19, label_col, 'Detention', bold=True)
            set_cell(19, value_col, float(trip.detention))
            set_cell(20, label_col, 'Additional Stop', bold=True)
            set_cell(20, value_col, 0)
            set_cell(21, label_col, 'Total Freight', bold=True)
            set_cell(21, value_col, float(trip.freight))
            set_cell(23, label_col, 'Departure Date', bold=True)
            set_cell(23, value_col, trip.trip_date)
            set_cell(24, label_col, 'Arrival Date', bold=True)
            set_cell(24, value_col, trip.trip_date)
            set_cell(25, label_col, 'Delivery', bold=True)
            set_cell(25, value_col, trip.trip_date)
        else:
            set_cell(13, label_col, 'Trip #', bold=True)
            set_cell(14, label_col, 'Bilty #', bold=True)
            set_cell(15, label_col, 'Client', bold=True)
            set_cell(16, label_col, 'Route', bold=True)
            set_cell(18, label_col, 'Freight', bold=True)
            set_cell(19, label_col, 'Detention', bold=True)
            set_cell(20, label_col, 'Additional Stop', bold=True)
            set_cell(21, label_col, 'Total Freight', bold=True)
            set_cell(23, label_col, 'Departure Date', bold=True)
            set_cell(24, label_col, 'Arrival Date', bold=True)
            set_cell(25, label_col, 'Delivery', bold=True)

    for column in range(3, 9):
        ws.column_dimensions[get_column_letter(column)].width = 18

    for row in range(1, 51):
        ws.row_dimensions[row].height = 18
    ws.row_dimensions[1].height = 45

    logo_path = Path(__file__).resolve().parent.parent / "static" / "images" / "amglogo.png"
    if logo_path.exists():
        try:
            logo_image = XLImage(str(logo_path))
            logo_image.width = 170
            logo_image.height = 55
            ws.add_image(logo_image, 'A1')
            set_cell(2, 1, 'Al Murad Logistics', bold=True)
        except Exception:
            pass

    ws.sheet_view.showGridLines = False

    set_cell(44, 5, 'Total', bold=True)
    set_cell(44, 6, 0)
    set_cell(44, 7, 'Total', bold=True)
    set_cell(44, 8, 0)
    set_cell(46, 7, 'Trip Expense', bold=True)
    set_cell(46, 8, 0)
    set_cell(47, 7, 'Fuel expense', bold=True)
    set_cell(47, 8, 0)
    set_cell(48, 7, 'Major Maintenance', bold=True)
    set_cell(48, 8, 0)
    set_cell(49, 7, 'Total Expenses', bold=True)
    set_cell(49, 8, 0)

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="trip_sheet.xlsx"'
    wb.save(response)
    return response


@login_required
def trips_pdf(request):
    report_context = _get_report_context(request)
    trips = list(report_context["trips"].order_by("trip_date")[:3])

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        rightMargin=30,
        leftMargin=30,
        topMargin=30,
        bottomMargin=18,
    )

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        "Title",
        parent=styles["Heading1"],
        fontSize=18,
        leading=22,
        spaceAfter=6,
    )
    normal = styles["Normal"]
    cell_style = ParagraphStyle(
        "Cell",
        parent=styles["BodyText"],
        fontSize=8,
        leading=10,
        spaceAfter=0,
        spaceBefore=0,
    )

    logo_path = Path(__file__).resolve().parent.parent / "static" / "images" / "amglogo.png"
    header_elements = []
    if logo_path.exists():
        try:
            logo = RLImage(str(logo_path), width=150, height=60)
            header_elements.append(logo)
        except Exception:
            pass

    company_para = Paragraph(
        "<b>Al Murad Logistics</b><br/>"
        "<font size=14>Trip Sheet Own</font><br/>"
        f"<font size=10>Generated on {dt.date.today().strftime('%d %b, %Y')}</font>",
        normal,
    )
    header_data = [[header_elements[0] if header_elements else "", company_para]]
    header_table = Table(header_data, colWidths=[160, 360])
    header_table.setStyle(TableStyle([
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("ALIGN", (1, 0), (1, 0), "LEFT"),
        ("LEFTPADDING", (0, 0), (-1, -1), 0),
        ("RIGHTPADDING", (0, 0), (-1, -1), 0),
        ("TOPPADDING", (0, 0), (-1, -1), 0),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 0),
    ]))

    elements = [header_table, Spacer(1, 12)]

    if trips:
        first_trip = trips[0]
        info_data = [
            ["Job #", str(first_trip.job.job_number)],
            ["Vehicle Number", first_trip.vehicle.vehicle_number],
            ["Driver 1 Name", first_trip.vehicle.driver.name if first_trip.vehicle.driver else ""],
            ["Driver 2 Name", ""],
            ["Meter Out", ""],
            ["Meter In", ""],
            ["KMs", ""],
            ["Fuel in Liters", ""],
            ["Fuel Average", ""],
        ]
    else:
        info_data = [["No OWN trips found", ""]]

    info_table = Table(info_data, colWidths=[140, 320], hAlign="LEFT")
    info_table.setStyle(
        TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.whitesmoke),
            ("TEXTCOLOR", (0, 0), (-1, -1), colors.black),
            ("ALIGN", (0, 0), (0, -1), "LEFT"),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("INNERGRID", (0, 0), (-1, -1), 0.25, colors.grey),
            ("BOX", (0, 0), (-1, -1), 0.5, colors.black),
        ])
    )
    elements.append(info_table)
    elements.append(Spacer(1, 14))

    trip_header = [
        "Trip #",
        "Bilty #",
        "Client",
        "Route",
        "Freight",
        "Detention",
        "Total Freight",
        "Date",
    ]
    trip_rows = [trip_header]

    for trip in trips:
        client_name = trip.client.name if trip.client else ""
        route_text = str(trip.route)
        trip_rows.append([
            Paragraph(str(trip.trip_no), cell_style),
            Paragraph(str(trip.bilty_number), cell_style),
            Paragraph(client_name, cell_style),
            Paragraph(route_text, cell_style),
            Paragraph(f"{trip.freight:.0f}", cell_style),
            Paragraph(f"{trip.detention:.0f}", cell_style),
            Paragraph(f"{trip.freight:.0f}", cell_style),
            Paragraph(trip.trip_date.strftime("%d %b, %Y"), cell_style),
        ])

    trip_table = Table(trip_rows, repeatRows=1, colWidths=[35, 55, 150, 120, 45, 45, 50, 50])
    trip_table.setStyle(
        TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.lightblue),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("ALIGN", (0, 0), (-1, 0), "CENTER"),
            ("ALIGN", (0, 1), (1, -1), "LEFT"),
            ("ALIGN", (2, 1), (3, -1), "LEFT"),
            ("ALIGN", (4, 1), (7, -1), "RIGHT"),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("GRID", (0, 0), (-1, -1), 0.25, colors.grey),
            ("FONTSIZE", (0, 0), (-1, -1), 8),
            ("BOTTOMPADDING", (0, 0), (-1, 0), 6),
            ("TOPPADDING", (0, 0), (-1, 0), 6),
        ])
    )
    elements.append(trip_table)

    doc.build(elements)
    pdf_value = buffer.getvalue()
    buffer.close()

    response = HttpResponse(pdf_value, content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="trip_sheet_own.pdf"'
    return response


def trip_add(request):
    form = TripForm(request.POST or None)
    if request.method == "POST" and form.is_valid():
        form.save()
        return redirect("trip_list")
    return render(request, "operations/trip_form.html", {"form": form})

def trip_edit(request, trip_id):
    trip = get_object_or_404(Trip, id=trip_id)
    form = TripForm(request.POST or None, instance=trip)
    if request.method == "POST" and form.is_valid():
        form.save()
        return redirect("trip_list")
    return render(request, "operations/trip_form.html", {"form": form, "trip": trip})

def trip_delete(request, trip_id):
    trip = get_object_or_404(Trip, id=trip_id)
    trip.delete()
    return redirect("trip_list")

# ================= MAINTENANCE =================
def maintenance_list(request):
    records = VehicleMaintenance.objects.all().order_by("-change_date")
    return render(request, "maintenance/maintenance_list.html", {"records": records})

def maintenance_add(request):
    vehicles = Vehicle.objects.all()
    if request.method == "POST":
        try:
            km = int(request.POST.get("change_km") or 0)
        except:
            km = 0
        VehicleMaintenance.objects.create(
            vehicle_id=request.POST.get("vehicle"),
            maintenance_type=request.POST.get("maintenance_type"),
            change_date=request.POST.get("change_date"),
            change_km=km,
            next_due_date=request.POST.get("next_due_date") or None,
            remarks=request.POST.get("remarks") or "",
        )
        return redirect("maintenance_list")
    return render(request, "maintenance/maintenance_form.html", {"vehicles": vehicles})

def job_view_trips(request, job_id):
    job = get_object_or_404(Job, id=job_id)
    trips = Trip.objects.filter(job=job).order_by("-trip_date")
    return render(request, "operations/trip_list.html", {"trips": trips, "job": job})

def job_invoice_pdf(request, job_id):
    # Job fetch karein (id ki bajaye job_number use karein jo primary key hai)
    job = get_object_or_404(Job, job_number=job_id)

    # Is job se linked saari trips fetch karein
    trips = job.trips.select_related("client", "route").order_by("trip_date")

    # Har trip ke sath uska fuel/expense data attach karein taake template mein
    # dobara query na karni pare
    fuel_rows = []
    for trip in trips:
        fuel_rows.extend(trip.expenses.all())

    expense_totals = Expense.objects.filter(trip__job=job).aggregate(
        toll_tax=Sum("toll_tax"),
        inam=Sum("inam"),
        police=Sum("police"),
        food=Sum("food"),
        card=Sum("card"),
        maintenance=Sum("maintenance"),
        other=Sum("other"),
        fuel_liter=Sum("fuel_liter"),
        fuel_amount=Sum("fuel_amount"),
    )
    for key, value in expense_totals.items():
        expense_totals[key] = float(value or 0)

    trip_expense_total = (
        expense_totals["toll_tax"] + expense_totals["inam"] + expense_totals["police"]
        + expense_totals["food"] + expense_totals["card"] + expense_totals["maintenance"]
        + expense_totals["other"]
    )
    fuel_expense_total = expense_totals["fuel_amount"]
    total_expenses = trip_expense_total + fuel_expense_total

    total_kms = sum(trip.route.distance_km for trip in trips)
    fuel_average = round(total_kms / expense_totals["fuel_liter"], 2) if expense_totals["fuel_liter"] else 0

    total_freight = sum(float(trip.freight) for trip in trips)
    net_profit = total_freight - total_expenses
    nr_percent = round((net_profit / total_freight) * 100, 1) if total_freight else 0

    context = {
        'job': job,
        'vehicle': job.vehicle,
        'trips': trips,
        'fuel_rows': fuel_rows,
        'total_kms': total_kms,
        'fuel_liters': expense_totals["fuel_liter"],
        'fuel_average': fuel_average,
        'expense_totals': expense_totals,
        'trip_expense_total': trip_expense_total,
        'fuel_expense_total': fuel_expense_total,
        'total_expenses': total_expenses,
        'total_freight': total_freight,
        'net_profit': net_profit,
        'nr_percent': nr_percent,
        'invoice_date': date.today(),
    }

    template_path = 'operations/invoice_pdf.html'
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="Trip_Sheet_Job_{job_id}.pdf"'

    template = get_template(template_path)
    html = template.render(context)

    pisa_status = pisa.CreatePDF(html, dest=response)

    if pisa_status.err:
        return HttpResponse('PDF banane mein masla hua', status=400)
    return response