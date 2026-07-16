from openpyxl import load_workbook
import json
import datetime
from pathlib import Path

path = Path('AMG.xlsx')
wb = load_workbook(path, data_only=True)
info = []
for sheet in wb.sheetnames:
    ws = wb[sheet]
    rows = []
    for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if i > 12:
            break
        rowvals = []
        for x in row:
            if isinstance(x, (datetime.datetime, datetime.date)):
                rowvals.append(x.isoformat())
            elif isinstance(x, datetime.time):
                rowvals.append(x.strftime('%H:%M:%S'))
            else:
                rowvals.append(x if x is not None else '')
        rows.append(rowvals)
    info.append({'sheet': sheet, 'rows': rows, 'shape': [ws.max_row, ws.max_column]})

with open('amg_summary.json', 'w', encoding='utf-8') as f:
    json.dump(info, f, ensure_ascii=False, indent=2)

print('wrote amg_summary.json')
